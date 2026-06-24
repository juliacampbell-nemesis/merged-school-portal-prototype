#!/usr/bin/env python3
"""Generate roles-matrix.html FROM roles-matrix.xlsx (single source of truth).

The xlsx is the source of truth (the deliverable Alongside/TC review). This script
renders it to a browsable, Helix-styled HTML page with a "Download .xlsx" button
that links to the canonical file. Run after any xlsx change:

    python3 gen-roles-matrix-html.py

Editorial prose not present in the xlsx (TC-mapping explainer, footer) lives here.
"""
import html
import re
import openpyxl

SRC = "roles-matrix.xlsx"
OUT = "roles-matrix.html"


def esc(v):
    return html.escape("" if v is None else str(v))


def cells_for(sheet, row, role_cols, style_map, default_classes):
    out = []
    for ci in role_cols:
        val = sheet.cell(row, ci).value
        val = "" if val is None else str(val).strip()
        cls, bold = style_map.get(val, ("", None))
        classes = ["text-center"]
        if cls:
            classes.append(cls)
            if bold:
                classes.append("font-bold")
        else:
            classes.extend(default_classes)
        if ci == 6:  # Clinical Staff column
            classes.append("role-clinical")
        out.append(f'<td class="{" ".join(classes)}">{esc(val)}</td>')
    return "".join(out)


def wrap_parens(text):
    return re.sub(r"(\([^)]*\))", r'<span class="text-xs text-tc-gray-400">\1</span>', esc(text))


# ---- editorial prose (not in the xlsx) ----
TC_MAPPING_NOTE = (
    "TC's existing PHI flag becomes the <em>split signal</em> between Clinical Staff "
    "and School Staff. Migration: any current School Staff with PHI = ON moves to "
    "Clinical Staff automatically."
)
LEGEND_PHI_NOTE = "PHI = Protected Health Info  ·  Defaults shown; school admin can override"
FOOTER_NOTE = "Throwaway analysis · not a spec · generated from roles-matrix.xlsx"

wb = openpyxl.load_workbook(SRC, data_only=True)

# ---------- Legend (symbols + source note) ----------
lg = wb["Legend"]
legend_items, source_note = [], ""
for row in lg.iter_rows(values_only=True):
    if not row:
        continue
    sym, meaning = (row[0], row[1] if len(row) > 1 else "")
    if sym in ("✓", "●", "—", "+"):
        legend_items.append((sym, str(meaning or "").split(" (")[0]))
    elif sym == "Source":
        source_note = str(meaning or "")

SYM_CLS = {"✓": "cell-yes", "●": "cell-cfg", "—": "cell-no", "+": "cell-new"}
legend_html = "\n".join(
    f'      <span class="inline-flex items-center gap-2"><span class="size-4 rounded-sm '
    f'{SYM_CLS.get(s, "")} inline-flex items-center justify-center font-bold">{esc(s)}</span> {esc(m)}</span>'
    for s, m in legend_items
)

# ---------- Main matrix ----------
rm = wb["Roles matrix"]
MATRIX_STYLE = {"✓": ("cell-yes", True), "●": ("cell-cfg", True), "—": ("cell-no", False)}
matrix_rows = []
for r in range(2, rm.max_row + 1):
    a, b = rm.cell(r, 1).value, rm.cell(r, 2).value
    if b in (None, ""):  # merged group-header banner
        label = str(a or "")
        if " — " in label:
            main, suf = label.split(" — ", 1)
            inner = f'{esc(main)} <span class="text-tc-gray-400 font-normal normal-case ml-2">— {esc(suf)}</span>'
        else:
            inner = esc(label)
        matrix_rows.append(
            f'      <tr class="section-header"><td colspan="7" class="px-4 py-2 text-xs uppercase tracking-wide">{inner}</td></tr>'
        )
    else:
        cells = cells_for(rm, r, range(3, 8), MATRIX_STYLE, [])
        note = esc(rm.cell(r, 8).value)
        matrix_rows.append(
            f'      <tr><td class="px-4 py-2.5 font-medium">{esc(b)}</td>{cells}'
            f'<td class="px-4 py-2 text-xs text-tc-gray-700">{note}</td></tr>'
        )
matrix_html = "\n".join(matrix_rows)

# ---------- Composable attributes ----------
ca = wb["Composable attributes"]
CA_STYLE = {"ON": ("cell-yes", True), "OFF": ("cell-no", False),
            "Optional": ("cell-cfg", True), "Configurable": ("cell-cfg", True),
            "All 5": ("cell-yes", True)}
ca_rows = []
for r in range(2, ca.max_row + 1):
    attr = ca.cell(r, 1).value
    if attr in (None, ""):
        continue
    cells = cells_for(ca, r, range(2, 7), CA_STYLE, ["text-xs"])
    gates = esc(ca.cell(r, 7).value)
    ca_rows.append(
        f'        <tr><td class="px-4 py-2.5 font-medium">{esc(attr)}</td>{cells}'
        f'<td class="px-4 py-2 text-xs text-tc-gray-700">{gates}</td></tr>'
    )
ca_html = "\n".join(ca_rows)

# ---------- Role mappings (two cards by source product) ----------
mp = wb["Role mappings"]
groups = {}
for r in range(2, mp.max_row + 1):
    prod = mp.cell(r, 1).value
    if prod in (None, ""):
        continue
    groups.setdefault(str(prod), []).append((mp.cell(r, 2).value, mp.cell(r, 3).value))


def mapping_rows(pairs):
    out = []
    for src, merged in pairs:
        merged_txt = wrap_parens(merged)
        clinical = "Clinical Staff" in str(merged)
        if clinical:
            merged_txt = merged_txt.replace("Clinical Staff", '<span class="text-tc-navy-700">Clinical Staff</span>')
            out.append(
                f'          <tr class="bg-tc-navy-50"><td class="py-2 text-tc-gray-700 px-2">{wrap_parens(src)}</td>'
                f'<td class="py-2 font-semibold text-right px-2">→ {merged_txt}</td></tr>'
            )
        else:
            out.append(
                f'          <tr><td class="py-2 text-tc-gray-700">{wrap_parens(src)}</td>'
                f'<td class="py-2 font-semibold text-right">→ {merged_txt}</td></tr>'
            )
    return "\n".join(out)


alongside_html = mapping_rows(groups.get("Alongside", []))
tc_html = mapping_rows(groups.get("TimelyCare", []))

# ---------- Open questions ----------
oq = wb["Open questions"]
oq_rows = []
for r in range(2, oq.max_row + 1):
    q = oq.cell(r, 2).value
    if q in (None, ""):
        continue
    q = str(q)
    if ". " in q:
        head, rest = q.split(". ", 1)
        oq_rows.append(f'      <li><strong class="text-tc-gray-900">{esc(head)}.</strong> {esc(rest)}</li>')
    else:
        oq_rows.append(f"      <li>{esc(q)}</li>")
oq_html = "\n".join(oq_rows)

# ---------- Assemble ----------
doc = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width,initial-scale=1" />
<title>Merged School Portal — Roles Matrix (analysis)</title>
<link rel="stylesheet" href="https://use.typekit.net/ayf2bms.css">
<script src="https://cdn.tailwindcss.com"></script>
<script src="https://unpkg.com/lucide@latest/dist/umd/lucide.js"></script>
<script>
  tailwind.config = {{
    theme: {{ extend: {{
      fontFamily: {{ sans: ['adelle-sans', 'ui-sans-serif', 'system-ui', 'sans-serif'] }},
      colors: {{
        'tc-navy': {{ 50: '#E5EDF4', 100: '#C2D6E5', 500: '#16538E', 700: '#0F3D6B', 900: '#0A2B4A' }},
        'tc-sage': {{ 50: '#ECF4EC', 100: '#D4E6D4', 500: '#5C8F5C', 700: '#3F6B3F' }},
        'tc-orange': {{ 50: '#FCEFE3', 100: '#F8DCC2', 500: '#D17A3C', 700: '#A5582A' }},
        'tc-berry': {{ 50: '#F8E7EC', 100: '#EDC3CF', 500: '#B33A5C', 700: '#8A2A47' }},
        'tc-gray': {{ 50: '#F6F6F4', 100: '#ECECE7', 200: '#D9D9D2', 400: '#8C8C84', 700: '#3F3F3A', 900: '#1F1F1C' }},
        'tc-cream': '#FBF7EE',
      }},
    }} }},
  }};
</script>
<style>
  body {{ font-family: 'adelle-sans', ui-sans-serif, system-ui, sans-serif; }}
  .cell-yes {{ background-color: #ECF4EC; color: #3F6B3F; }}
  .cell-no  {{ background-color: #FBF7EE; color: #8C8C84; }}
  .cell-cfg {{ background-color: #FCEFE3; color: #A5582A; }}
  .cell-new {{ background-color: #E5EDF4; color: #0F3D6B; }}
  .role-clinical {{ background-color: #E5EDF4; }}
  .matrix th, .matrix td {{ border-bottom: 1px solid #ECECE7; }}
  .matrix tbody tr:hover td {{ background-color: #FBF7EE; }}
  .matrix tbody tr:hover td.role-clinical {{ background-color: #C2D6E5; }}
  .section-header td {{ background-color: #F6F6F4; font-weight: 700; color: #1F1F1C; }}
  @media print {{
    body {{ font-size: 10px; }}
    .no-print {{ display: none; }}
    .matrix {{ font-size: 10px; }}
    .page-break {{ page-break-before: always; }}
  }}
</style>
</head>
<body class="bg-tc-cream text-tc-gray-900">

<header class="border-b border-tc-gray-200 bg-white">
  <div class="max-w-[1400px] mx-auto px-6 py-5">
    <div class="flex items-baseline justify-between gap-4 flex-wrap">
      <div>
        <h1 class="text-2xl font-bold tracking-tight">Merged School Portal — Roles Matrix</h1>
        <p class="text-sm text-tc-gray-700 mt-1">5 user types · TimelyCare base + one new role (<strong class="text-tc-navy-700">Clinical Staff</strong>) absorbing Alongside's Counselor / Social Worker / School Psychologist.</p>
      </div>
      <div class="flex flex-col items-end gap-2 no-print">
        <a href="roles-matrix.xlsx" download class="inline-flex items-center gap-2 h-9 px-4 rounded-md bg-tc-navy-500 text-white text-sm font-semibold hover:bg-tc-navy-700 transition-colors"><i data-lucide="download" class="size-4"></i>Download spreadsheet (.xlsx)</a>
        <div class="text-xs text-tc-gray-400">{esc(source_note) or "Merged School Portal — Roles Matrix"} · for discussion, not built</div>
      </div>
    </div>

    <!-- Legend -->
    <div class="mt-4 flex items-center gap-4 flex-wrap text-xs">
{legend_html}
      <span class="text-tc-gray-400 ml-2">{LEGEND_PHI_NOTE}</span>
    </div>
  </div>
</header>

<main class="max-w-[1400px] mx-auto px-6 py-6">

  <!-- ============ MAIN MATRIX ============ -->
  <table class="matrix w-full text-sm bg-white rounded-lg shadow-sm overflow-hidden">
    <thead>
      <tr class="bg-tc-navy-500 text-white">
        <th class="text-left px-4 py-3 font-semibold sticky left-0 bg-tc-navy-500 z-10" style="width: 320px;">Capability</th>
        <th class="px-3 py-3 font-semibold text-center" style="width: 130px;">District<br>Administrator</th>
        <th class="px-3 py-3 font-semibold text-center" style="width: 130px;">District<br>Staff</th>
        <th class="px-3 py-3 font-semibold text-center" style="width: 130px;">School<br>Administrator</th>
        <th class="px-3 py-3 font-semibold text-center bg-tc-navy-700" style="width: 130px;">Clinical Staff<br><span class="text-[10px] font-normal opacity-90">+ NEW</span></th>
        <th class="px-3 py-3 font-semibold text-center" style="width: 130px;">School<br>Staff</th>
        <th class="px-4 py-3 text-left font-semibold" style="width: 280px;">Notes / origin</th>
      </tr>
    </thead>
    <tbody>
{matrix_html}
    </tbody>
  </table>

  <!-- ============ ATTRIBUTES & DEFAULTS ============ -->
  <section class="mt-8">
    <h2 class="text-base font-bold mb-3">Composable attributes (set per user, on top of role)</h2>
    <table class="matrix w-full text-sm bg-white rounded-lg shadow-sm overflow-hidden">
      <thead>
        <tr class="bg-tc-gray-100 text-tc-gray-900">
          <th class="text-left px-4 py-3 font-semibold" style="width: 280px;">Attribute</th>
          <th class="px-3 py-3 font-semibold text-center" style="width: 130px;">District Admin</th>
          <th class="px-3 py-3 font-semibold text-center" style="width: 130px;">District Staff</th>
          <th class="px-3 py-3 font-semibold text-center" style="width: 130px;">School Admin</th>
          <th class="px-3 py-3 font-semibold text-center bg-tc-navy-50" style="width: 130px;">Clinical Staff</th>
          <th class="px-3 py-3 font-semibold text-center" style="width: 130px;">School Staff</th>
          <th class="px-4 py-3 text-left font-semibold">What it gates</th>
        </tr>
      </thead>
      <tbody>
{ca_html}
      </tbody>
    </table>
  </section>

  <!-- ============ MAPPING REFERENCE ============ -->
  <section class="mt-8 grid grid-cols-2 gap-4">
    <div class="bg-white rounded-lg shadow-sm p-5 border border-tc-gray-200">
      <h3 class="text-base font-bold mb-3 flex items-center gap-2">
        <span class="size-2 rounded-full bg-tc-orange-500"></span>
        Alongside → merged role mapping
      </h3>
      <table class="w-full text-sm">
        <tbody class="divide-y divide-tc-gray-100">
{alongside_html}
        </tbody>
      </table>
    </div>

    <div class="bg-white rounded-lg shadow-sm p-5 border border-tc-gray-200">
      <h3 class="text-base font-bold mb-3 flex items-center gap-2">
        <span class="size-2 rounded-full bg-tc-navy-500"></span>
        TimelyCare → merged role mapping
      </h3>
      <table class="w-full text-sm">
        <tbody class="divide-y divide-tc-gray-100">
{tc_html}
        </tbody>
      </table>
      <p class="text-xs text-tc-gray-700 mt-3 leading-relaxed">
        {TC_MAPPING_NOTE}
      </p>
    </div>
  </section>

  <!-- ============ OPEN QUESTIONS ============ -->
  <section class="mt-8 bg-white rounded-lg shadow-sm p-5 border border-tc-gray-200">
    <h3 class="text-base font-bold mb-3">Open questions before this can ship</h3>
    <ol class="space-y-2 text-sm text-tc-gray-700 list-decimal pl-5">
{oq_html}
    </ol>
  </section>

  <footer class="mt-8 mb-12 text-xs text-tc-gray-400 flex items-center justify-between">
    <span>{FOOTER_NOTE}</span>
    <a href="merged-portal-prototype.html" class="text-tc-navy-500 hover:underline">← back to prototype</a>
  </footer>

</main>

<script>
  if (window.lucide && window.lucide.createIcons) window.lucide.createIcons();
</script>
</body>
</html>
"""

with open(OUT, "w") as f:
    f.write(doc)
print(f"wrote {OUT} ({len(doc)} bytes)")
